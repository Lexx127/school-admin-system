"""
Attendance Excel export router.
Requires: pip install openpyxl
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import User, StudentAttendance, Student, Class, UserRole, AttendanceStatus, SchoolCalendarDay, DayType
from auth import require_role
from datetime import date, timedelta
import io

router = APIRouter(prefix="/exports", tags=["exports"])
attendance_export_router = APIRouter(prefix="/attendance/export", tags=["exports"])

def hex_to_rgb(hex_color: str):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def _build_attendance_workbook(class_id: int, start_date: date, end_date: date, sheet_title: str, title_text: str, db: Session):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

    cls = db.query(Class).filter(Class.id == class_id).first()
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.user_id).all()

    records = db.query(StudentAttendance).filter(
        StudentAttendance.class_id == class_id,
        StudentAttendance.date >= start_date,
        StudentAttendance.date <= end_date
    ).all()

    record_map = {}
    for r in records:
        record_map[(r.student_id, r.date)] = r.status

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.merge_cells("A1:H1")
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal='center')

    day_count = (end_date - start_date).days + 1
    days = [start_date + timedelta(days=i) for i in range(day_count)]
    headers = ["#", "Student Name"] + [d.strftime("%a %d %b") for d in days] + ["Total Present"]
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal='center')

    fills = {
        AttendanceStatus.PRESENT: PatternFill("solid", fgColor="D1FAE5"),
        AttendanceStatus.ABSENT: PatternFill("solid", fgColor="FEE2E2"),
        AttendanceStatus.LATE: PatternFill("solid", fgColor="FEF9C3"),
    }
    holiday_fill = PatternFill("solid", fgColor="E5E7EB")
    calendar_days = db.query(SchoolCalendarDay).filter(
        SchoolCalendarDay.date >= start_date,
        SchoolCalendarDay.date <= end_date
    ).all()
    holiday_dates = {d.date for d in calendar_days if d.day_type != DayType.SCHOOL_DAY}

    for i, student in enumerate(students, 1):
        name = f"{student.user.first_name} {student.user.last_name}"
        row = [i, name]
        present_count = 0
        for d in days:
            if d in holiday_dates:
                row.append("H")
                continue
            status = record_map.get((student.id, d), None)
            if status == AttendanceStatus.PRESENT:
                present_count += 1
            row.append(status.value if status else "–")
        row.append(present_count)
        ws.append(row)
        row_idx = ws.max_row
        for col_idx, d in enumerate(days, start=3):
            if d in holiday_dates:
                ws.cell(row=row_idx, column=col_idx).fill = holiday_fill
                continue
            status = record_map.get((student.id, d), None)
            if status and status in fills:
                ws.cell(row=row_idx, column=col_idx).fill = fills[status]

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    return wb, cls

@router.get("/attendance/weekly")
def export_weekly_attendance(
    class_id: int = Query(...),
    week_start: date = Query(...),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.PRINCIPAL, UserRole.TEACHER)),
    db: Session = Depends(get_db)
):
    week_end = week_start + timedelta(days=4)
    cls = db.query(Class).filter(Class.id == class_id).first()
    wb, cls = _build_attendance_workbook(
        class_id=class_id,
        start_date=week_start,
        end_date=week_end,
        sheet_title="Weekly Attendance",
        title_text=f"Weekly Attendance - {cls.name if cls else class_id} - Week of {week_start.strftime('%d %b %Y')}",
        db=db
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{cls.name if cls else class_id}_{week_start.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/attendance/termly")
def export_termly_attendance(
    class_id: int = Query(...),
    term_start: date = Query(...),
    term_end: date = Query(...),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.PRINCIPAL, UserRole.TEACHER)),
    db: Session = Depends(get_db)
):
    if term_start > term_end:
        raise HTTPException(status_code=400, detail="term_start must be on or before term_end")

    cls = db.query(Class).filter(Class.id == class_id).first()
    wb, cls = _build_attendance_workbook(
        class_id=class_id,
        start_date=term_start,
        end_date=term_end,
        sheet_title="Termly Attendance",
        title_text=f"Termly Attendance - {cls.name if cls else class_id} - {term_start.strftime('%d %b %Y')} to {term_end.strftime('%d %b %Y')}",
        db=db
    )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"attendance_{cls.name if cls else class_id}_{term_start.isoformat()}_{term_end.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@attendance_export_router.get("/weekly")
def export_weekly_attendance_alias(
    class_id: int = Query(...),
    week_start: date = Query(...),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.PRINCIPAL, UserRole.TEACHER)),
    db: Session = Depends(get_db)
):
    return export_weekly_attendance(class_id=class_id, week_start=week_start, current_user=current_user, db=db)


@attendance_export_router.get("/termly")
def export_termly_attendance_alias(
    class_id: int = Query(...),
    term_start: date = Query(...),
    term_end: date = Query(...),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.PRINCIPAL, UserRole.TEACHER)),
    db: Session = Depends(get_db)
):
    return export_termly_attendance(
        class_id=class_id,
        term_start=term_start,
        term_end=term_end,
        current_user=current_user,
        db=db
    )
